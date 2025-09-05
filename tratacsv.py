import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

def abrecsv(dados, caminho):

    #df = pd.read_csv(caminho)
    #if "CLIENTE" in df.columns:
    #    df = df.drop(columns=["CLIENTE"])

    nova_linha = {"O.S.": dados["Id"]
                  ,"MATRICULA": dados["matricula"]
                  ,"LOTE E QD": dados["complemento"]
                  ,"ENDEREÇO": (dados["endereco"]+" Nº "+dados["numero"])
                  ,"CONSTRUTOR": dados["nome"]
                  ,"STATUS": (dados["grupo"]+dados["atividade"]+" ABERTURA: "+dados["data"])
                  ,"VALOR LAUDO" : 0} 

    try:
        wb = openpyxl.load_workbook(caminho)
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {caminho}")
        return
    
    ws = wb.active

    empty_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(cell.value in (None, "") for cell in row):
            empty_row_idx = i
            break
    
    values = list(nova_linha.values())
    if empty_row_idx:
        ws.insert_rows(empty_row_idx)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=empty_row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    else:
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(caminho)
